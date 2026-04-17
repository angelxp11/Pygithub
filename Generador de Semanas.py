import tkinter as tk
from tkinter import filedialog, ttk, messagebox
from openpyxl import load_workbook
import re
from pathlib import Path
import os

# =====================================================
# 🔧 UTILIDADES
# =====================================================

def convertir_google_drive_descarga(url):
    if "docs.google.com/document" in url:
        match = re.search(r"/d/([^/]+)", url)
        if match:
            file_id = match.group(1)
            return f"https://docs.google.com/document/d/{file_id}/export?format=pdf"
    return url


def convertir_youtube_embed(url):
    if not url:
        return None

    url = str(url)

    if "youtu.be" in url:
        video_id = url.split("/")[-1].split("?")[0]
        return f"https://www.youtube.com/embed/{video_id}"

    match = re.search(r"v=([^&]+)", url)
    if match:
        return f"https://www.youtube.com/embed/{match.group(1)}"

    match_shorts = re.search(r"shorts/([^?&/]+)", url)
    if match_shorts:
        return f"https://www.youtube.com/embed/{match_shorts.group(1)}"

    return None


def tarjeta_video_bloqueado(url):
    return f"""
    <div style="background:#fff3cd; padding:15px; border-radius:15px; margin:10px 0;">
        <p style="margin:0 0 8px 0; font-weight:bold;">⚠️ Video no disponible</p>
        <a href="{url}" target="_blank"
        style="padding:8px 15px; background:#ff4d4d; color:white; border-radius:10px; text-decoration:none;">
        ▶ Ver en YouTube
        </a>
    </div>
    """


# =====================================================
# 📚 GENERAR SEMANAS
# =====================================================

def generar_semanas(ws):

    colores = ["#e3f2fd", "#fff3cd", "#d4edda", "#f8d7da", "#e2e3ff"]
    rangos_cortes = {}

    def extraer_rango(texto, nombre_corte):
        numeros = re.findall(r'\d+', str(texto))
        if len(numeros) >= 2:
            inicio, fin = int(numeros[0]), int(numeros[1])
        elif len(numeros) == 1:
            inicio = fin = int(numeros[0])
        else:
            return

        for i in range(inicio, fin + 1):
            rangos_cortes[i] = nombre_corte

    # CORTES
    extraer_rango(ws["B9"].value, "🥇 Primer Corte")
    extraer_rango(ws["B10"].value, "🥈 Segundo Corte")
    extraer_rango(ws["B11"].value, "🥉 Tercer Corte")

    exam_columns = {
        "🥇 Primer Corte": 2,
        "🥈 Segundo Corte": 3,
        "🥉 Tercer Corte": 4
    }

    # ÚLTIMA SEMANA POR CORTE
    last_weeks = {}
    for week, cut in rangos_cortes.items():
        if cut not in last_weeks or week > last_weeks[cut]:
            last_weeks[cut] = week

    semanas = []
    col = 2

    while ws.cell(row=2, column=col).value:
        semanas.append(col)
        col += 1

    html = ""

    for i in range(0, len(semanas), 4):

        html += """
<div style="display:flex; flex-wrap:wrap; gap:20px; justify-content:center; margin-bottom:30px;">
"""

        for j in range(i, min(i + 4, len(semanas))):

            col = semanas[j]
            numero_semana = j + 1
            color = colores[j % len(colores)]

            semana_nombre = ws.cell(row=2, column=col).value or ""
            tema = ws.cell(row=3, column=col).value or ""
            corte = rangos_cortes.get(numero_semana, "")

            contenido_html = ""

            for fila in range(4, ws.max_row + 1):

                tipo = ws.cell(row=fila, column=1).value
                contenido = ws.cell(row=fila, column=col).value

                if not tipo or not contenido:
                    continue

                tipo = str(tipo).strip().upper()
                contenido = str(contenido).strip()

                # VIDEO
                if tipo == "VIDEO EXPLICACIÓN":

                    embed = convertir_youtube_embed(contenido)

                    if embed:
                        contenido_html += f"""
                        <p><strong>🎥 Video:</strong></p>
                        <iframe src="{embed}" width="100%" height="180"
                        style="border-radius:12px;"></iframe>
                        """
                    else:
                        contenido_html += tarjeta_video_bloqueado(contenido)

                # GUIA
                elif tipo == "GUIA EXPLICACIÓN":

                    link = convertir_google_drive_descarga(contenido)

                    contenido_html += f"""
                    <a href="{link}" target="_blank"
                    style="display:inline-block;margin:8px 0;padding:10px 18px;
                    background:linear-gradient(135deg,#4CAF50,#2E7D32);
                    color:white;border-radius:12px;text-decoration:none;font-weight:bold;">
                    📘 Descargar Guía
                    </a>
                    """

                # TALLER
                elif tipo == "TAREA O TALLER":

                    link = convertir_google_drive_descarga(contenido)

                    contenido_html += f"""
                    <a href="{link}" target="_blank"
                    style="display:inline-block;margin:8px 0;padding:10px 18px;
                    background:linear-gradient(135deg,#2196F3,#1565C0);
                    color:white;border-radius:12px;text-decoration:none;font-weight:bold;">
                    📝 Descargar Taller
                    </a>
                    """

                # SUBIR
                elif tipo == "TALLERESSUBIR":

                    contenido_html += f"""
                    <a href="{contenido}" target="_blank"
                    style="display:inline-block;margin:8px 0;padding:12px 20px;
                    background:linear-gradient(135deg,#ff4d4d,#b71c1c);
                    color:white;border-radius:12px;text-decoration:none;font-weight:bold;">
                    ⬆ Subir Actividad
                    </a>
                    """

            # 🎯 EXAMEN (RESTAURADO)
            if numero_semana == last_weeks.get(corte, -1):
                col_exam = exam_columns.get(corte, None)
                if col_exam:
                    link = ws.cell(row=31, column=col_exam).value
                    if link:
                        contenido_html += f"""
                        <a href="{link}" target="_blank"
                        style="display:inline-block;margin:8px 0;padding:10px 18px;
                        background:linear-gradient(135deg,#FF9800,#E65100);
                        color:white;border-radius:12px;text-decoration:none;font-weight:bold;">
                        📊 EXAMEN {corte.split()[1]} Corte
                        </a>
                        """

            html += f"""
<details style="width:300px; background:{color}; padding:15px; border-radius:20px;
box-shadow:0 8px 18px rgba(0,0,0,0.08);">
<summary style="cursor:pointer; font-weight:bold;">🌟 {semana_nombre}</summary>
<p><strong>📖 Tema:</strong> {tema}</p>
<p><strong>📌 {corte}</strong></p>
{contenido_html}
</details>
"""

        html += "</div>"

    return html


# =====================================================
# 🖥️ INTERFAZ
# =====================================================

def cargar_archivo():
    ruta = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xlsm")])
    if ruta:
        entry_ruta.delete(0, tk.END)
        entry_ruta.insert(0, ruta)

        wb = load_workbook(ruta)
        combo_hojas["values"] = wb.sheetnames
        combo_hojas.current(0)


def ejecutar():
    ruta = entry_ruta.get()
    hoja = combo_hojas.get()

    if not ruta or not hoja:
        messagebox.showerror("Error", "Selecciona archivo y hoja")
        return

    try:
        wb = load_workbook(ruta)
        ws = wb[hoja]

        semanas_html = generar_semanas(ws)

        html_final = f"""
<div style="font-family: Comic Sans MS, cursive; max-width:1100px; margin:auto;">
  <details style="background:#ffeccf; padding:20px; border-radius:20px;">
    <summary style="cursor:pointer; font-size:22px; font-weight:bold;">
      📚 Semanas de Trabajo
    </summary>
    <div style="margin-top:20px;">
      {semanas_html}
    </div>
  </details>
</div>
"""

        salida = Path(ruta).parent / "semanas_generadas.html"

        with open(salida, "w", encoding="utf-8") as f:
            f.write(html_final)

        messagebox.showinfo("Éxito", f"Archivo generado:\n{salida}")

        os.startfile(Path(ruta).parent)

    except Exception as e:
        messagebox.showerror("Error", str(e))


# =====================================================
# 🚀 VENTANA
# =====================================================

root = tk.Tk()
root.title("Generador de Semanas PRO")
root.geometry("520x260")

tk.Label(root, text="Archivo Excel:").pack(pady=5)
entry_ruta = tk.Entry(root, width=55)
entry_ruta.pack()

tk.Button(root, text="📂 Buscar archivo", command=cargar_archivo).pack(pady=5)

tk.Label(root, text="Seleccionar hoja:").pack(pady=5)
combo_hojas = ttk.Combobox(root, width=40)
combo_hojas.pack()

tk.Button(root, text="🚀 Generar HTML", command=ejecutar).pack(pady=15)

root.mainloop()