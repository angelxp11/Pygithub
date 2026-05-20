import tkinter as tk
from tkinter import filedialog, ttk, messagebox
from openpyxl import load_workbook
import os
import subprocess


def generar_html_malla(ruta_malla, hoja_malla):

    wb_malla = load_workbook(ruta_malla)
    ws_malla = wb_malla[hoja_malla]

    # =====================================================
    # 📘 GENERAR CONTENIDO
    # =====================================================

    malla_contenido = ""
    titulo_actual = None
    identificador_actual = None

    for fila in range(1, ws_malla.max_row + 1):

        texto = ws_malla.cell(row=fila, column=1).value
        tipo = ws_malla.cell(row=fila, column=2).value

        if not texto or not tipo:
            continue

        texto = str(texto).strip()
        tipo = str(tipo).strip().lower()

        # =====================================================
        # 📌 TITULOS
        # =====================================================

        if tipo.startswith("titulo"):

            if titulo_actual:
                malla_contenido += """
        </div>
      </details>
"""

            titulo_actual = texto
            identificador_actual = tipo.replace("titulo", "")

            malla_contenido += f"""
      <details
        style="background: #ffffff; padding: 18px; border-radius: 20px; margin: 15px 0; box-shadow: 0 6px 18px rgba(0,0,0,0.08);">
        <summary style="cursor: pointer; font-size: 20px; font-weight: bold;">📌
          <strong>{titulo_actual}</strong></summary>
        <div style="margin-top: 15px;">
"""

        # =====================================================
        # ✨ CONTENIDOS
        # =====================================================

        elif tipo.startswith("contenido"):

            numero_contenido = tipo.replace("contenido", "")

            if numero_contenido == identificador_actual:

                malla_contenido += f"""
          <div style="margin: 8px 0; padding-left: 20px; font-size: 15px;">✨
            {texto}</div>
"""

    # Cerrar último bloque
    if titulo_actual:
        malla_contenido += """
        </div>
      </details>
"""

    # =====================================================
    # 🌐 HTML FINAL
    # =====================================================

    html = f"""
<div
  style="font-family: Comic Sans MS, cursive; max-width: 1100px; margin: auto;">

  <details
    style="margin-bottom: 25px; background: linear-gradient(135deg,#d9f2ff,#ffffff); padding: 25px; border-radius: 25px; box-shadow: 0 8px 20px rgba(0,0,0,0.1);">

    <summary style="cursor: pointer; font-size: 22px; font-weight: bold;">
      📘 Malla Curricular
    </summary>

    <div style="margin-top: 15px;">

{malla_contenido}

    </div>

  </details>

</div>
"""

    nombre_archivo = "malla_generada.html"

    with open(nombre_archivo, "w", encoding="utf-8") as f:
        f.write(html)

    return os.path.abspath(nombre_archivo)


# =====================================================
# 🖥️ INTERFAZ
# =====================================================

ruta_malla = ""


def seleccionar_malla():

    global ruta_malla

    ruta_malla = filedialog.askopenfilename(
        filetypes=[("Excel", "*.xlsx")]
    )

    if ruta_malla:
        cargar_hojas(ruta_malla)


def cargar_hojas(ruta):

    wb = load_workbook(ruta, read_only=True)

    combo_malla["values"] = wb.sheetnames

    if wb.sheetnames:
        combo_malla.current(0)


def generar():

    if not ruta_malla:
        messagebox.showerror("Error", "Selecciona un archivo Excel")
        return

    hoja_malla = combo_malla.get()

    archivo = generar_html_malla(
        ruta_malla,
        hoja_malla
    )

    messagebox.showinfo(
        "Éxito",
        "HTML generado correctamente"
    )

    # Abrir automáticamente en VS Code
    try:
        subprocess.Popen(["code", archivo])

    except:
        os.startfile(archivo)


# =====================================================
# 🚀 VENTANA
# =====================================================

root = tk.Tk()

root.title("Generador de Malla HTML")
root.geometry("520x300")
root.resizable(False, False)

titulo = tk.Label(
    root,
    text="📘 Generador de Malla Curricular",
    font=("Arial", 16, "bold")
)

titulo.pack(pady=20)

btn_excel = tk.Button(
    root,
    text="Seleccionar archivo Excel",
    command=seleccionar_malla,
    width=30,
    height=2
)

btn_excel.pack(pady=10)

combo_malla = ttk.Combobox(
    root,
    width=55
)

combo_malla.pack(pady=5)

btn_generar = tk.Button(
    root,
    text="🚀 Generar HTML",
    command=generar,
    bg="#4CAF50",
    fg="white",
    font=("Arial", 12, "bold"),
    width=25,
    height=2
)

btn_generar.pack(pady=30)

root.mainloop()